"""Analyze and extract appointment data from AI Appointments Set 2025.xlsx"""
import openpyxl
import csv
import json
from datetime import datetime

INPUT_FILE = r'C:\Users\peter\OneDrive\Documents\Yes Right Pty Ltd\AI Appointments Set 2025.xlsx'
OUTPUT_CSV = r'C:\Users\peter\Downloads\CC\CRM\All_Appointments_Extracted.csv'


def extract_appointments():
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    all_appointments = []

    for sheet_name in wb.sheetnames:
        print(f"\n=== Processing: {sheet_name} ===")
        ws = wb[sheet_name]

        # Read all rows
        rows = list(ws.iter_rows(values_only=True, max_row=500))
        if not rows:
            continue

        # The data appears to be transposed (field names in first column, data in subsequent columns)
        # Let's check the structure

        # Find row indices for key fields
        field_indices = {}
        for i, row in enumerate(rows):
            if row and row[0]:
                field_name = str(row[0]).lower().strip()
                if 'company' in field_name:
                    field_indices['company'] = i
                elif 'appt date' in field_name or 'appointment date' in field_name or field_name == 'date':
                    field_indices['date'] = i
                elif 'appt time' in field_name or 'time' in field_name:
                    field_indices['time'] = i
                elif 'email' in field_name:
                    field_indices['email'] = i
                elif 'status' in field_name:
                    field_indices['status'] = i
                elif 'quality' in field_name:
                    field_indices['quality'] = i
                elif 'follow' in field_name:
                    field_indices['followup'] = i
                elif 'phone' in field_name or 'mobile' in field_name:
                    field_indices['phone'] = i
                elif 'name' in field_name and 'company' not in field_name:
                    field_indices['name'] = i
                elif 'retell' in field_name or 'log' in field_name:
                    field_indices['retell_log'] = i

        print(f"  Found fields: {list(field_indices.keys())}")

        if not field_indices:
            print(f"  Skipping - no recognizable fields")
            continue

        # Determine number of records (columns after first)
        max_cols = max(len(row) for row in rows if row)
        print(f"  Max columns: {max_cols}")

        # Extract each appointment (each column after the first)
        for col_idx in range(1, max_cols):
            appt = {
                'source_sheet': sheet_name,
                'company': None,
                'date': None,
                'time': None,
                'email': None,
                'status': None,
                'quality': None,
                'followup': None,
                'phone': None,
                'name': None,
                'retell_log': None,
            }

            for field, row_idx in field_indices.items():
                if row_idx < len(rows) and col_idx < len(rows[row_idx]):
                    value = rows[row_idx][col_idx]
                    if value is not None:
                        # Handle datetime objects
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value).strip()
                        if value and value.lower() not in ['none', 'nan', '']:
                            appt[field] = value

            # Only keep if has meaningful data
            if appt.get('company') or appt.get('email') or appt.get('date'):
                all_appointments.append(appt)

    wb.close()
    return all_appointments


def main():
    print("=" * 60)
    print("EXTRACTING APPOINTMENT DATA")
    print("=" * 60)

    appointments = extract_appointments()

    print(f"\n{'='*60}")
    print(f"EXTRACTION RESULTS")
    print(f"{'='*60}")
    print(f"Total appointments found: {len(appointments)}")

    # Stats
    has_email = sum(1 for a in appointments if a.get('email'))
    has_company = sum(1 for a in appointments if a.get('company'))
    has_date = sum(1 for a in appointments if a.get('date'))
    has_status = sum(1 for a in appointments if a.get('status'))

    print(f"\nData coverage:")
    print(f"  Has Email:   {has_email}")
    print(f"  Has Company: {has_company}")
    print(f"  Has Date:    {has_date}")
    print(f"  Has Status:  {has_status}")

    # Status breakdown
    statuses = {}
    for a in appointments:
        s = a.get('status', 'Unknown')
        if s:
            statuses[s] = statuses.get(s, 0) + 1

    print(f"\nStatus breakdown:")
    for status, count in sorted(statuses.items(), key=lambda x: -x[1])[:15]:
        print(f"  {status[:40]:40} {count}")

    # Write CSV
    if appointments:
        print(f"\nWriting to: {OUTPUT_CSV}")

        fieldnames = ['source_sheet', 'company', 'name', 'email', 'phone',
                      'date', 'time', 'status', 'quality', 'followup', 'retell_log']

        with open(OUTPUT_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(appointments)

        print("Done!")

    # Print sample
    print(f"\n{'='*60}")
    print("SAMPLE APPOINTMENTS (first 10)")
    print(f"{'='*60}")
    for i, appt in enumerate(appointments[:10]):
        print(f"\n{i+1}. {appt.get('company', 'N/A')}")
        print(f"   Email: {appt.get('email', 'N/A')}")
        print(f"   Date: {appt.get('date', 'N/A')} {appt.get('time', '')}")
        print(f"   Status: {appt.get('status', 'N/A')}")

    return appointments


if __name__ == "__main__":
    appointments = main()
